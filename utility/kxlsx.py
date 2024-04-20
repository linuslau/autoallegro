# Note: openpyxl version can't be greater than 3.0.10, it will crash with filter active

from utility.klogger import *
import openpyxl as xl
import pandas as pd

klogger = KLogger()
logger = klogger.getlogger()
class KXlsx(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = xl.load_workbook(self.file_name)
        self.df = pd.read_excel(self.file_name, None)
        pass

    def __del__(self):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def get_all_sheets_wb(self):
        logger.info(self.wb.sheetnames)
        return self.wb.sheetnames

    def get_all_sheets_pd(self):
        logger.info(self.df.keys())
        return self.df.keys()

    def get_all_sheets(self, type='wb'):
        if type == 'wb':
            return self.get_all_sheets_wb()
        elif type == 'pd':
            return self.get_all_sheets_pd()
        else:
            return self.get_all_sheets_pd()

    def get_active_sheet(self):
        # logger.info("Active sheet: ", wb.active)  # will crash
        print("Active sheet: ", self.wb.active)
        return self.wb.active

    def set_active_sheet(self):
        self.wb._active_sheet_index = 4
        #logger.info("Active sheet: ", wb.active) # will crash
        print("Active sheet: ", self.wb.active)
        return self.wb.active

if __name__ == '__main__':
    file = r'c:\Users\kliu59\Documents\Code\Allegro\autoallegro\Netlist comparision table.xlsx'
    kxlsx = KXlsx(file)
    sheetnames = kxlsx.get_all_sheets('wb')
    sheetnames = kxlsx.get_all_sheets('pd')
    sheetnames = kxlsx.get_all_sheets()
    pass